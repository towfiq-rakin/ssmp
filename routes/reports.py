"""
Report Generation Module
Handles Excel and PDF export for students, scholarships, and stipends
"""
from flask import Blueprint, send_file, flash, redirect, url_for
from flask_login import login_required, current_user
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from datetime import datetime
import io
from models import User, AcademicRecord, Scholarship, Stipend, Department, Admin
from extensions import db

reports_bp = Blueprint('reports', __name__, url_prefix='/admin')


# ==================== STUDENT REPORTS ====================

@reports_bp.route('/reports/students/excel')
@login_required
def export_students_excel():
    """Export students list to Excel"""
    if not isinstance(current_user, Admin):
        flash('Access denied', 'danger')
        return redirect(url_for('main.home'))
    
    # Get students from admin's department
    students = db.session.query(User, AcademicRecord).join(
        AcademicRecord, User.student_id == AcademicRecord.student_id
    ).filter(User.dept_id == current_user.dept_id).all()
    
    department = current_user.get_department()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = f"Student Report - {department.name}"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Subtitle
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Student ID', 'Name', 'Email', 'CGPA', 'Sem 1', 'Sem 2', 'Sem 3', 'Sem 4']
    ws.append([])  # Empty row
    ws.append(headers)
    
    # Style headers
    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    for student, academic_record in students:
        ws.append([
            student.student_id,
            student.name,
            student.email,
            round(academic_record.cgpa, 2) if academic_record.cgpa else 'N/A',
            round(academic_record.semester_1_gpa, 2) if academic_record.semester_1_gpa else 'N/A',
            round(academic_record.semester_2_gpa, 2) if academic_record.semester_2_gpa else 'N/A',
            round(academic_record.semester_3_gpa, 2) if academic_record.semester_3_gpa else 'N/A',
            round(academic_record.semester_4_gpa, 2) if academic_record.semester_4_gpa else 'N/A'
        ])
    
    # Style data rows
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    for col in ['D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 12
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Students_{department.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@reports_bp.route('/reports/students/pdf')
@login_required
def export_students_pdf():
    """Export students list to PDF"""
    if not isinstance(current_user, Admin):
        flash('Access denied', 'danger')
        return redirect(url_for('main.home'))
    
    # Get students from admin's department
    students = db.session.query(User, AcademicRecord).join(
        AcademicRecord, User.student_id == AcademicRecord.student_id
    ).filter(User.dept_id == current_user.dept_id).all()
    
    department = current_user.get_department()
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1976d2'),
        spaceAfter=12,
        alignment=1  # Center
    )
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontSize=8,
        leading=10
    )
    
    # Title
    title = Paragraph(f"<b>Student Report - {department.name}</b>", title_style)
    elements.append(title)
    
    # Subtitle
    subtitle = Paragraph(
        f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
        styles['Normal']
    )
    elements.append(subtitle)
    elements.append(Spacer(1, 20))
    
    # Table data with Paragraph for text wrapping
    data = [[
        Paragraph('<b>Student ID</b>', cell_style),
        Paragraph('<b>Name</b>', cell_style),
        Paragraph('<b>Email</b>', cell_style),
        Paragraph('<b>CGPA</b>', cell_style)
    ]]
    
    for student, academic_record in students:
        data.append([
            Paragraph(str(student.student_id), cell_style),
            Paragraph(student.name, cell_style),
            Paragraph(student.email, cell_style),
            Paragraph(f"{academic_record.cgpa:.2f}" if academic_record.cgpa else 'N/A', cell_style)
        ])
    
    # Create table with adjusted column widths for portrait
    table = Table(data, colWidths=[1.2*inch, 2.0*inch, 2.5*inch, 0.8*inch])
    
    # Table style
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"Students_{department.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )


# ==================== SCHOLARSHIP REPORTS ====================

@reports_bp.route('/reports/scholarships/excel')
@login_required
def export_scholarships_excel():
    """Export awarded scholarships to Excel"""
    if not isinstance(current_user, Admin):
        flash('Access denied', 'danger')
        return redirect(url_for('main.home'))
    
    # Get scholarships from admin's department
    scholarships = db.session.query(Scholarship, User).join(
        User, Scholarship.student_id == User.student_id
    ).filter(User.dept_id == current_user.dept_id).order_by(Scholarship.awarded_at.desc()).all()
    
    department = current_user.get_department()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Scholarships"
    
    # Header styling
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = f"Scholarship Awards Report - {department.name}"
    ws['A1'].font = Font(bold=True, size=16, color="2C5F2D")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Subtitle
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Summary
    total_amount = sum([s.amount for s, u in scholarships])
    ws.merge_cells('A3:G3')
    ws['A3'] = f"Total Scholarships: {len(scholarships)} | Total Amount: ৳{total_amount:,.2f}"
    ws['A3'].font = Font(bold=True, size=11)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Student ID', 'Student Name', 'Type', 'Amount', 'Semester', 'Awarded Date', 'Awarded Time']
    ws.append([])  # Empty row
    ws.append(headers)
    
    # Style headers
    for cell in ws[5]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    for scholarship, student in scholarships:
        ws.append([
            scholarship.student_id,
            scholarship.student_name,
            scholarship.type,
            scholarship.amount,
            scholarship.semester,
            scholarship.awarded_at.strftime('%Y-%m-%d'),
            scholarship.awarded_at.strftime('%I:%M %p')
        ])
    
    # Style data rows
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # Format amount column
        row[3].number_format = '৳#,##0.00'
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Scholarships_{department.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@reports_bp.route('/reports/scholarships/pdf')
@login_required
def export_scholarships_pdf():
    """Export awarded scholarships to PDF"""
    if not isinstance(current_user, Admin):
        flash('Access denied', 'danger')
        return redirect(url_for('main.home'))
    
    # Get scholarships from admin's department
    scholarships = db.session.query(Scholarship, User).join(
        User, Scholarship.student_id == User.student_id
    ).filter(User.dept_id == current_user.dept_id).order_by(Scholarship.awarded_at.desc()).all()
    
    department = current_user.get_department()
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#2c5f2d'),
        spaceAfter=12,
        alignment=1
    )
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontSize=8,
        leading=10
    )
    
    # Title
    title = Paragraph(f"<b>Scholarship Awards Report - {department.name}</b>", title_style)
    elements.append(title)
    
    # Subtitle
    subtitle = Paragraph(
        f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
        styles['Normal']
    )
    elements.append(subtitle)
    
    # Summary
    total_amount = sum([s.amount for s, u in scholarships])
    summary = Paragraph(
        f"<b>Total Scholarships:</b> {len(scholarships)} | <b>Total Amount:</b> {total_amount:,.2f}",
        styles['Normal']
    )
    elements.append(summary)
    elements.append(Spacer(1, 20))
    
    # Table data with Paragraph for text wrapping
    data = [[
        Paragraph('<b>Student ID</b>', cell_style),
        Paragraph('<b>Name</b>', cell_style),
        Paragraph('<b>Type</b>', cell_style),
        Paragraph('<b>Amount</b>', cell_style),
        Paragraph('<b>Semester</b>', cell_style),
        Paragraph('<b>Awarded Date</b>', cell_style)
    ]]
    
    for scholarship, student in scholarships:
        data.append([
            Paragraph(str(scholarship.student_id), cell_style),
            Paragraph(scholarship.student_name, cell_style),
            Paragraph(scholarship.type, cell_style),
            Paragraph(f"{scholarship.amount:,.2f}", cell_style),
            Paragraph(scholarship.semester, cell_style),
            Paragraph(scholarship.awarded_at.strftime('%Y-%m-%d'), cell_style)
        ])
    
    # Create table with adjusted column widths for portrait
    table = Table(data, colWidths=[0.8*inch, 1.4*inch, 1.4*inch, 0.9*inch, 0.9*inch, 0.9*inch])
    
    # Table style
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"Scholarships_{department.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )


# ==================== STIPEND REPORTS ====================

@reports_bp.route('/reports/stipends/excel')
@login_required
def export_stipends_excel():
    """Export awarded stipends to Excel"""
    if not isinstance(current_user, Admin):
        flash('Access denied', 'danger')
        return redirect(url_for('main.home'))
    
    # Get stipends from admin's department
    stipends = db.session.query(Stipend, User).join(
        User, Stipend.student_id == User.student_id
    ).filter(User.dept_id == current_user.dept_id).order_by(Stipend.awarded_at.desc()).all()
    
    department = current_user.get_department()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Stipends"
    
    # Header styling
    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = f"Stipend Awards Report - {department.name}"
    ws['A1'].font = Font(bold=True, size=16, color="1565C0")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Subtitle
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Summary
    total_amount = sum([s.amount for s, u in stipends])
    ws.merge_cells('A3:G3')
    ws['A3'] = f"Total Stipends: {len(stipends)} | Total Amount: {total_amount:,.2f}"
    ws['A3'].font = Font(bold=True, size=11)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Student ID', 'Student Name', 'Type', 'Amount', 'Semester', 'Awarded Date', 'Awarded Time']
    ws.append([])  # Empty row
    ws.append(headers)
    
    # Style headers
    for cell in ws[5]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    for stipend, student in stipends:
        ws.append([
            stipend.student_id,
            stipend.student_name,
            stipend.type,
            stipend.amount,
            stipend.semester,
            stipend.awarded_at.strftime('%Y-%m-%d'),
            stipend.awarded_at.strftime('%I:%M %p')
        ])
    
    # Style data rows
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # Format amount column
        row[3].number_format = '#,##0.00'
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Stipends_{department.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@reports_bp.route('/reports/stipends/pdf')
@login_required
def export_stipends_pdf():
    """Export awarded stipends to PDF"""
    if not isinstance(current_user, Admin):
        flash('Access denied', 'danger')
        return redirect(url_for('main.home'))
    
    # Get stipends from admin's department
    stipends = db.session.query(Stipend, User).join(
        User, Stipend.student_id == User.student_id
    ).filter(User.dept_id == current_user.dept_id).order_by(Stipend.awarded_at.desc()).all()
    
    department = current_user.get_department()
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1976d2'),
        spaceAfter=12,
        alignment=1
    )
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontSize=8,
        leading=10
    )
    
    # Title
    title = Paragraph(f"<b>Stipend Awards Report - {department.name}</b>", title_style)
    elements.append(title)
    
    # Subtitle
    subtitle = Paragraph(
        f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
        styles['Normal']
    )
    elements.append(subtitle)
    
    # Summary
    total_amount = sum([s.amount for s, u in stipends])
    summary = Paragraph(
        f"<b>Total Stipends:</b> {len(stipends)} | <b>Total Amount:</b> ৳{total_amount:,.2f}",
        styles['Normal']
    )
    elements.append(summary)
    elements.append(Spacer(1, 20))
    
    # Table data with Paragraph for text wrapping
    data = [[
        Paragraph('<b>Student ID</b>', cell_style),
        Paragraph('<b>Name</b>', cell_style),
        Paragraph('<b>Type</b>', cell_style),
        Paragraph('<b>Amount</b>', cell_style),
        Paragraph('<b>Semester</b>', cell_style),
        Paragraph('<b>Awarded Date</b>', cell_style)
    ]]
    
    for stipend, student in stipends:
        data.append([
            Paragraph(str(stipend.student_id), cell_style),
            Paragraph(stipend.student_name, cell_style),
            Paragraph(stipend.type, cell_style),
            Paragraph(f"{stipend.amount:,.2f}", cell_style),
            Paragraph(stipend.semester, cell_style),
            Paragraph(stipend.awarded_at.strftime('%Y-%m-%d'), cell_style)
        ])
    
    # Create table with adjusted column widths for portrait
    table = Table(data, colWidths=[0.8*inch, 1.4*inch, 1.4*inch, 0.9*inch, 0.9*inch, 0.9*inch])
    
    # Table style
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2196F3')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"Stipends_{department.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )
